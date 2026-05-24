"""
将Neo4j图数据导出为多种格式（JSON、CSV、Excel、Cypher）
"""
import json
import csv
import os
import sys
from datetime import datetime
from typing import Dict, List, Any, Tuple

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("警告: pandas未安装，Excel导出功能不可用。请运行: pip install pandas openpyxl")

# 添加项目根目录到 Python 路径
# 这样可以从 reports/ 目录运行脚本
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from core.neo4j_store import Neo4jGraphStore


def export_to_json(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出图数据为JSON格式
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"graph_export_{timestamp}.json"
    
    print(f"\n开始导出图数据到: {output_file}")
    print("=" * 60)
    
    # 1. 导出所有节点
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    RETURN n.id as id, 
           n.name as name, 
           labels(n) as labels,
           n.is_concept as is_concept,
           n.entity_type as entity_type,
           properties(n) as properties
    """
    nodes = graph_store.query(nodes_query)

    # 补充“本体归属”信息（用于节点工作表）
    instance_type_query = """
    MATCH (n)-[r]->(c)
    WHERE type(r) IN ['类型', 'INSTANCE_OF']
    RETURN n.id as node_id, collect(DISTINCT c.name) as concept_names
    """
    instance_type_rows = graph_store.query(instance_type_query)
    instance_of_map = {}
    for row in instance_type_rows:
        nid = row.get("node_id", "")
        names = row.get("concept_names", []) or []
        if nid:
            instance_of_map[nid] = sorted([str(x) for x in names if x])

    concept_parent_query = """
    MATCH (child)-[r]->(parent)
    WHERE type(r) IN ['子类', 'SUB_CLASS_OF']
    RETURN child.id as child_id, collect(DISTINCT parent.name) as parent_names
    """
    concept_parent_rows = graph_store.query(concept_parent_query)
    concept_parent_map = {}
    for row in concept_parent_rows:
        cid = row.get("child_id", "")
        names = row.get("parent_names", []) or []
        if cid:
            concept_parent_map[cid] = sorted([str(x) for x in names if x])

    concept_child_query = """
    MATCH (child)-[r]->(parent)
    WHERE type(r) IN ['子类', 'SUB_CLASS_OF']
    RETURN parent.id as parent_id, collect(DISTINCT child.name) as child_names
    """
    concept_child_rows = graph_store.query(concept_child_query)
    concept_child_map = {}
    for row in concept_child_rows:
        pid = row.get("parent_id", "")
        names = row.get("child_names", []) or []
        if pid:
            concept_child_map[pid] = sorted([str(x) for x in names if x])

    # 补充“本体归属”信息
    instance_type_query = """
    MATCH (n)-[r]->(c)
    WHERE type(r) IN ['类型', 'INSTANCE_OF']
    RETURN n.id as node_id, collect(DISTINCT c.name) as concept_names
    """
    instance_type_rows = graph_store.query(instance_type_query)
    instance_of_map = {}
    for row in instance_type_rows:
        nid = row.get("node_id", "")
        names = row.get("concept_names", []) or []
        if nid:
            instance_of_map[nid] = sorted([str(x) for x in names if x])

    concept_parent_query = """
    MATCH (child)-[r]->(parent)
    WHERE type(r) IN ['子类', 'SUB_CLASS_OF']
    RETURN child.id as child_id, collect(DISTINCT parent.name) as parent_names
    """
    concept_parent_rows = graph_store.query(concept_parent_query)
    concept_parent_map = {}
    for row in concept_parent_rows:
        cid = row.get("child_id", "")
        names = row.get("parent_names", []) or []
        if cid:
            concept_parent_map[cid] = sorted([str(x) for x in names if x])

    concept_child_query = """
    MATCH (child)-[r]->(parent)
    WHERE type(r) IN ['子类', 'SUB_CLASS_OF']
    RETURN parent.id as parent_id, collect(DISTINCT child.name) as child_names
    """
    concept_child_rows = graph_store.query(concept_child_query)
    concept_child_map = {}
    for row in concept_child_rows:
        pid = row.get("parent_id", "")
        names = row.get("child_names", []) or []
        if pid:
            concept_child_map[pid] = sorted([str(x) for x in names if x])
    nodes_list = []
    concept_count = 0
    
    for record in nodes:
        node_data = {
            "id": record.get("id"),
            "name": record.get("name"),
            "labels": record.get("labels", []),
            "is_concept": record.get("is_concept", False),
            "entity_type": record.get("entity_type", ""),
            "properties": {k: v for k, v in (record.get("properties") or {}).items() 
                          if k not in ["id", "name", "is_concept", "entity_type"]}
        }
        nodes_list.append(node_data)
        if record.get("is_concept"):
            concept_count += 1
    
    print(f"  导出了 {len(nodes_list)} 个节点 (其中概念节点: {concept_count})")
    
    # 2. 导出所有关系
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           s.name as source_name,
           type(r) as relation_type,
           o.id as target_id,
           o.name as target_name,
           r.is_ontology as is_ontology,
           properties(r) as properties
    """
    relationships = graph_store.query(relationships_query)
    relationships_list = []
    ontology_rel_count = 0
    
    for record in relationships:
        rel_data = {
            "source": {
                "id": record.get("source_id"),
                "name": record.get("source_name")
            },
            "target": {
                "id": record.get("target_id"),
                "name": record.get("target_name")
            },
            "relation": record.get("relation_type"),
            "is_ontology": record.get("is_ontology", False),
            "properties": {k: v for k, v in (record.get("properties") or {}).items()}
        }
        relationships_list.append(rel_data)
        if record.get("is_ontology"):
            ontology_rel_count += 1
    
    print(f"  导出了 {len(relationships_list)} 条关系 (其中本体关系: {ontology_rel_count})")
    
    # 3. 获取统计信息
    print("正在获取统计信息...")
    stats = graph_store.get_stats()
    
    # 4. 组装导出数据
    export_data = {
        "metadata": {
            "export_time": datetime.now().isoformat(),
            "nodes_count": len(nodes_list),
            "concept_nodes_count": concept_count,
            "relationships_count": len(relationships_list),
            "ontology_relations_count": ontology_rel_count,
            "database_stats": stats
        },
        "nodes": nodes_list,
        "relationships": relationships_list
    }
    
    # 5. 写入JSON文件
    print(f"正在写入文件: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    print(f"  节点数: {len(nodes_list)} (概念: {concept_count})")
    print(f"  关系数: {len(relationships_list)} (本体: {ontology_rel_count})")
    
    return output_file


def export_simplified_json(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出简化版JSON（仅包含基本信息，文件更小）
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"graph_export_simple_{timestamp}.json"
    
    print(f"\n开始导出简化版图数据到: {output_file}")
    print("=" * 60)
    
    # 导出节点（仅ID和名称）
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    RETURN DISTINCT n.id as id, n.name as name, n.is_concept as is_concept
    """
    nodes = graph_store.query(nodes_query)
    nodes_dict = {record.get("id"): {
        "name": record.get("name"),
        "is_concept": record.get("is_concept", False)
    } for record in nodes if record.get("id")}
    
    concept_count = sum(1 for n in nodes_dict.values() if n.get("is_concept"))
    print(f"  导出了 {len(nodes_dict)} 个节点 (其中概念节点: {concept_count})")
    
    # 导出关系（简化版）
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           type(r) as relation_type,
           o.id as target_id,
           r.is_ontology as is_ontology,
           r.evidence as evidence,
           r.source_paper as source_paper
    """
    relationships = graph_store.query(relationships_query)
    relationships_list = []
    ontology_rel_count = 0
    
    for record in relationships:
        rel_data = {
            "source_id": record.get("source_id"),
            "target_id": record.get("target_id"),
            "relation": record.get("relation_type"),
            "is_ontology": record.get("is_ontology", False),
            "evidence": record.get("evidence", "")[:200] if record.get("evidence") else "",  # 限制长度
            "source_paper": record.get("source_paper", "")
        }
        relationships_list.append(rel_data)
        if record.get("is_ontology"):
            ontology_rel_count += 1
    
    print(f"  导出了 {len(relationships_list)} 条关系 (其中本体关系: {ontology_rel_count})")
    
    # 组装数据
    export_data = {
        "metadata": {
            "export_time": datetime.now().isoformat(),
            "nodes_count": len(nodes_dict),
            "concept_nodes_count": concept_count,
            "relationships_count": len(relationships_list),
            "ontology_relations_count": ontology_rel_count
        },
        "nodes": nodes_dict,
        "relationships": relationships_list
    }
    
    # 写入文件
    print(f"正在写入文件: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    
    return output_file


def export_cypher_format(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出为Cypher格式（可以用于重新导入）
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"graph_export_{timestamp}.cypher"
    
    print(f"\n开始导出Cypher格式到: {output_file}")
    print("=" * 60)
    
    cypher_statements = []
    
    # 导出节点
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    RETURN n.id as id, 
           n.name as name, 
           labels(n) as labels,
           n.is_concept as is_concept,
           n.entity_type as entity_type
    """
    nodes = graph_store.query(nodes_query)
    node_count = 0
    concept_count = 0
    
    for record in nodes:
        node_id = record.get("id", "").replace("'", "\\'")
        node_name = record.get("name", "").replace("'", "\\'")
        labels = record.get("labels", ["Entity"])
        is_concept = record.get("is_concept", False)
        entity_type = record.get("entity_type", "")
        
        # 构建标签字符串
        label_str = ":".join(labels)
        
        # 构建属性字符串
        props = f"id: '{node_id}', name: '{node_name}'"
        if is_concept:
            props += f", is_concept: true"
        if entity_type:
            props += f", entity_type: '{entity_type}'"
        
        cypher = f"CREATE (n:{label_str} {{{props}}});"
        cypher_statements.append(cypher)
        
        node_count += 1
        if is_concept:
            concept_count += 1
    
    print(f"  导出了 {node_count} 个节点 (其中概念节点: {concept_count})")
    
    # 导出关系
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           type(r) as relation_type,
           o.id as target_id,
           r.is_ontology as is_ontology,
           r.evidence as evidence,
           r.source_paper as source_paper,
           r.relation_original as relation_original
    """
    relationships = graph_store.query(relationships_query)
    rel_count = 0
    ontology_rel_count = 0
    
    for record in relationships:
        source_id = record.get("source_id", "").replace("'", "\\'")
        target_id = record.get("target_id", "").replace("'", "\\'")
        rel_type = record.get("relation_type", "RELATES_TO")
        is_ontology = record.get("is_ontology", False)
        evidence = (record.get("evidence", "") or "").replace("'", "\\'")[:100]
        source_paper = (record.get("source_paper", "") or "").replace("'", "\\'")
        relation_original = (record.get("relation_original", "") or "").replace("'", "\\'")
        
        # 构建关系属性
        props = []
        if evidence:
            props.append(f"evidence: '{evidence}'")
        if source_paper:
            props.append(f"source_paper: '{source_paper}'")
        if relation_original:
            props.append(f"relation_original: '{relation_original}'")
        if is_ontology:
            props.append(f"is_ontology: true")
        
        props_str = ", ".join(props) if props else ""
        
        cypher = f"""MATCH (s {{id: '{source_id}'}}), (o {{id: '{target_id}'}})
CREATE (s)-[r:{rel_type} {{{props_str}}}]->(o);"""
        cypher_statements.append(cypher)
        
        rel_count += 1
        if is_ontology:
            ontology_rel_count += 1
    
    print(f"  导出了 {rel_count} 条关系 (其中本体关系: {ontology_rel_count})")
    
    # 写入文件
    print(f"正在写入文件: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("// Neo4j Graph Export\n")
        f.write(f"// Export time: {datetime.now().isoformat()}\n")
        f.write(f"// Nodes: {node_count} (Concepts: {concept_count}), Relationships: {rel_count} (Ontology: {ontology_rel_count})\n\n")
        f.write("\n".join(cypher_statements))
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    
    return output_file


def export_to_csv(graph_store: Neo4jGraphStore, output_dir: str = None):
    """
    导出图数据为CSV格式（节点和关系分别导出为两个CSV文件）
    
    Args:
        graph_store: Neo4j图存储实例
        output_dir: 输出目录（可选，默认为当前目录）
    """
    if output_dir is None:
        output_dir = "."
    else:
        os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nodes_file = os.path.join(output_dir, f"nodes_{timestamp}.csv")
    relationships_file = os.path.join(output_dir, f"relationships_{timestamp}.csv")
    
    print(f"\n开始导出CSV格式图数据")
    print("=" * 60)
    
    # 1. 导出节点
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    WHERE n:Entity OR n:Ontology
    RETURN n.id as id, 
           n.name as name, 
           labels(n) as labels,
           n.is_concept as is_concept,
           n.entity_type as entity_type,
           n.created_at as created_at,
           n.updated_at as updated_at
    ORDER BY n.id
    """
    nodes = graph_store.query(nodes_query)
    
    node_count = 0
    concept_count = 0
    with open(nodes_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        # 写入表头
        writer.writerow([
            'id', 'name', 'labels', 'is_concept', 'entity_type', 
            'created_at', 'updated_at'
        ])
        
        for record in nodes:
            node_id = record.get("id", "")
            node_name = record.get("name", "")
            labels = record.get("labels", [])
            labels_str = ";".join(labels) if labels else ""
            is_concept = record.get("is_concept", False)
            entity_type = record.get("entity_type", "")
            created_at = record.get("created_at", "")
            updated_at = record.get("updated_at", "")
            
            # 处理时间戳
            if created_at:
                try:
                    created_at = datetime.fromtimestamp(created_at / 1000).isoformat()
                except:
                    created_at = str(created_at)
            if updated_at:
                try:
                    updated_at = datetime.fromtimestamp(updated_at / 1000).isoformat()
                except:
                    updated_at = str(updated_at)
            
            writer.writerow([
                node_id, node_name, labels_str, is_concept, entity_type,
                created_at, updated_at
            ])
            node_count += 1
            if is_concept:
                concept_count += 1
    
    print(f"  导出了 {node_count} 个节点 (概念: {concept_count}) 到: {nodes_file}")
    
    # 2. 导出关系
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           s.name as source_name,
           type(r) as relation_type,
           o.id as target_id,
           o.name as target_name,
           r.is_ontology as is_ontology,
           r.evidence as evidence,
           r.source_paper_id as source_paper_id,
           r.source_paper as source_paper
    ORDER BY s.id, o.id
    """
    relationships = graph_store.query(relationships_query)
    
    rel_count = 0
    ontology_rel_count = 0
    with open(relationships_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        # 写入表头
        writer.writerow([
            'source_id', 'source_name', 'target_id', 'target_name',
            'relation_type', 'is_ontology', 'evidence', 'source_paper_id', 'source_paper'
        ])
        
        for record in relationships:
            source_id = record.get("source_id", "")
            source_name = record.get("source_name", "")
            target_id = record.get("target_id", "")
            target_name = record.get("target_name", "")
            relation_type = record.get("relation_type", "")
            is_ontology = record.get("is_ontology", False)
            evidence = record.get("evidence", "")
            source_paper_id = record.get("source_paper_id", "")
            source_paper = record.get("source_paper", "")
            
            # 限制 evidence 长度（CSV 中避免过长）
            if evidence and len(evidence) > 500:
                evidence = evidence[:500] + "..."
            
            writer.writerow([
                source_id, source_name, target_id, target_name,
                relation_type, is_ontology, evidence, source_paper_id, source_paper
            ])
            rel_count += 1
            if is_ontology:
                ontology_rel_count += 1
    
    print(f"  导出了 {rel_count} 条关系 (本体: {ontology_rel_count}) 到: {relationships_file}")
    
    # 3. 计算文件大小
    nodes_size = os.path.getsize(nodes_file) / (1024 * 1024)  # MB
    rels_size = os.path.getsize(relationships_file) / (1024 * 1024)  # MB
    
    print(f"\n导出完成！")
    print(f"  节点文件: {nodes_file} ({nodes_size:.2f} MB)")
    print(f"  关系文件: {relationships_file} ({rels_size:.2f} MB)")
    print(f"  节点数: {node_count} (概念: {concept_count})")
    print(f"  关系数: {rel_count} (本体: {ontology_rel_count})")
    
    return nodes_file, relationships_file


def export_ontology_to_csv(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出本体层次结构为CSV
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"ontology_hierarchy_{timestamp}.csv"
    
    print(f"\n开始导出本体层次结构到: {output_file}")
    print("=" * 60)
    
    # 方法1: 查询显式的本体关系
    query_1 = """
    MATCH (s)-[r]->(o)
    WHERE r.is_ontology = true 
       OR type(r) IN ['SUB_CLASS_OF', 'INSTANCE_OF', 'RDF_TYPE', 'TYPE']
    RETURN s.id as subject_id,
           s.name as subject_name,
           labels(s) as subject_labels,
           s.is_concept as subject_is_concept,
           type(r) as relation,
           o.id as object_id,
           o.name as object_name,
           labels(o) as object_labels,
           o.is_concept as object_is_concept,
           r.evidence as evidence
    ORDER BY s.name, o.name
    """
    results = graph_store.query(query_1)
    
    if not results:
        print("  未找到显式本体关系，尝试方法2...")
        
        # 方法2: 查询所有可能的本体关系（基于概念标记）
        query_2 = """
        MATCH (s)-[r]->(o)
        WHERE (s.is_concept = true OR o.is_concept = true)
          AND NOT type(r) IN ['INSTANCE_OF']  // 排除明显的实例关系
        RETURN s.id as subject_id,
               s.name as subject_name,
               labels(s) as subject_labels,
               s.is_concept as subject_is_concept,
               type(r) as relation,
               o.id as object_id,
               o.name as object_name,
               labels(o) as object_labels,
               o.is_concept as object_is_concept,
               r.evidence as evidence
        ORDER BY s.name, o.name
        """
        results = graph_store.query(query_2)
    
    count = 0
    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        # 写入表头
        writer.writerow([
            'subject_id', 'subject_name', 'subject_is_concept', 
            'relation', 
            'object_id', 'object_name', 'object_is_concept',
            'evidence'
        ])
        
        for record in results:
            writer.writerow([
                record.get("subject_id", ""),
                record.get("subject_name", ""),
                record.get("subject_is_concept", False),
                record.get("relation", ""),
                record.get("object_id", ""),
                record.get("object_name", ""),
                record.get("object_is_concept", False),
                record.get("evidence", "")
            ])
            count += 1
    
    # 方法3: 如果有 Concept 标记的节点，导出所有概念节点
    if count == 0:
        print("  未找到本体关系，尝试方法3 - 导出概念节点...")
        query_3 = """
        MATCH (n)
        WHERE n.is_concept = true OR 'Concept' IN labels(n)
        RETURN n.id as concept_id,
               n.name as concept_name,
               labels(n) as labels,
               n.entity_type as entity_type
        ORDER BY n.name
        """
        concepts = graph_store.query(query_3)
        
        if concepts:
            with open(output_file, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['concept_id', 'concept_name', 'labels', 'entity_type'])
                for record in concepts:
                    writer.writerow([
                        record.get("concept_id", ""),
                        record.get("concept_name", ""),
                        ";".join(record.get("labels", [])),
                        record.get("entity_type", "")
                    ])
                count = len(concepts)
    
    file_size = os.path.getsize(output_file) / 1024  # KB
    print(f"  导出了 {count} 条本体记录")
    print(f"  文件: {output_file} ({file_size:.2f} KB)")
    
    return output_file


def export_ontology_excel(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    专门导出本体数据为Excel格式（包含概念节点、本体关系两个工作表）
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("pandas和openpyxl未安装，无法导出Excel。请运行: pip install pandas openpyxl")
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"ontology_export_{timestamp}.xlsx"
    
    print(f"\n开始导出本体数据到: {output_file}")
    print("=" * 60)
    
    # 1. 导出概念节点
    print("正在导出概念节点...")
    concepts_query = """
    MATCH (n)
    WHERE n.is_concept = true OR 'Concept' IN labels(n)
    RETURN n.id as id,
           n.name as name,
           labels(n) as labels,
           n.entity_type as entity_type,
           n.created_at as created_at
    ORDER BY n.name
    """
    concepts = graph_store.query(concepts_query)
    
    concepts_data = []
    for record in concepts:
        created_at = record.get("created_at", "")
        if created_at:
            try:
                created_at = datetime.fromtimestamp(created_at / 1000).strftime("%Y-%m-%d %H:%M:%S")
            except:
                created_at = str(created_at)
        
        concepts_data.append({
            'id': record.get("id", ""),
            'name': record.get("name", ""),
            'labels': ";".join(record.get("labels", [])),
            'entity_type': record.get("entity_type", ""),
            'created_at': created_at
        })
    
    df_concepts = pd.DataFrame(concepts_data)
    print(f"  导出了 {len(df_concepts)} 个概念节点")
    
    # 2. 导出本体关系
    print("正在导出本体关系...")
    relations_query = """
    MATCH (s)-[r]->(o)
    WHERE r.is_ontology = true 
       OR type(r) IN ['SUB_CLASS_OF', 'INSTANCE_OF', '为', '藏', '主', '生', '克', 
                      '合', '开窍于', '其华在', '走', '入', '伤', '胜', '畏', '恶']
    RETURN s.id as source_id,
           s.name as source_name,
           s.is_concept as source_is_concept,
           type(r) as relation_type,
           o.id as target_id,
           o.name as target_name,
           o.is_concept as target_is_concept,
           r.evidence as evidence,
           r.source_paper as source_paper
    ORDER BY source_name, target_name
    """
    relations = graph_store.query(relations_query)
    
    relations_data = []
    for record in relations:
        evidence = record.get("evidence", "")
        if evidence and len(evidence) > 500:
            evidence = evidence[:500] + "..."
        
        relations_data.append({
            'source_id': record.get("source_id", ""),
            'source_name': record.get("source_name", ""),
            'source_is_concept': record.get("source_is_concept", False),
            'relation_type': record.get("relation_type", ""),
            'target_id': record.get("target_id", ""),
            'target_name': record.get("target_name", ""),
            'target_is_concept': record.get("target_is_concept", False),
            'evidence': evidence,
            'source_paper': record.get("source_paper", "")
        })
    
    df_relations = pd.DataFrame(relations_data)
    print(f"  导出了 {len(df_relations)} 条本体关系")
    
    # 3. 写入Excel文件
    print(f"正在写入Excel文件: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_concepts.to_excel(writer, sheet_name='概念节点', index=False)
        df_relations.to_excel(writer, sheet_name='本体关系', index=False)
        
        # 添加统计信息
        stats_data = {
            '指标': ['概念节点数', '本体关系数', '导出时间'],
            '数值': [len(df_concepts), len(df_relations), datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        }
        df_stats = pd.DataFrame(stats_data)
        df_stats.to_excel(writer, sheet_name='统计信息', index=False)
    
    # 调整列宽
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    
    for sheet_name in ['概念节点', '本体关系', '统计信息']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    print(f"  概念节点: {len(df_concepts)}")
    print(f"  本体关系: {len(df_relations)}")
    
    return output_file


def export_to_excel(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出图数据为Excel格式（包含多个工作表：节点、关系、本体层次）
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("pandas和openpyxl未安装，无法导出Excel。请运行: pip install pandas openpyxl")
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"graph_export_{timestamp}.xlsx"
    
    print(f"\n开始导出Excel格式图数据到: {output_file}")
    print("=" * 60)
    
    # 1. 导出节点
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    WHERE n:Entity OR n:Ontology
    RETURN n.id as id, 
           n.name as name, 
           labels(n) as labels,
           n.is_concept as is_concept,
           n.entity_type as entity_type,
           n.created_at as created_at,
           n.updated_at as updated_at
    ORDER BY n.id
    """
    nodes = graph_store.query(nodes_query)

    # 补充“本体归属”信息（用于节点工作表）
    instance_type_query = """
    MATCH (n)-[r]->(c)
    WHERE type(r) IN ['类型', 'INSTANCE_OF']
    RETURN n.id as node_id, collect(DISTINCT c.name) as concept_names
    """
    instance_type_rows = graph_store.query(instance_type_query)
    instance_of_map = {}
    for row in instance_type_rows:
        nid = row.get("node_id", "")
        names = row.get("concept_names", []) or []
        if nid:
            instance_of_map[nid] = sorted([str(x) for x in names if x])

    concept_parent_query = """
    MATCH (child)-[r]->(parent)
    WHERE type(r) IN ['子类', 'SUB_CLASS_OF']
    RETURN child.id as child_id, collect(DISTINCT parent.name) as parent_names
    """
    concept_parent_rows = graph_store.query(concept_parent_query)
    concept_parent_map = {}
    for row in concept_parent_rows:
        cid = row.get("child_id", "")
        names = row.get("parent_names", []) or []
        if cid:
            concept_parent_map[cid] = sorted([str(x) for x in names if x])

    concept_child_query = """
    MATCH (child)-[r]->(parent)
    WHERE type(r) IN ['子类', 'SUB_CLASS_OF']
    RETURN parent.id as parent_id, collect(DISTINCT child.name) as child_names
    """
    concept_child_rows = graph_store.query(concept_child_query)
    concept_child_map = {}
    for row in concept_child_rows:
        pid = row.get("parent_id", "")
        names = row.get("child_names", []) or []
        if pid:
            concept_child_map[pid] = sorted([str(x) for x in names if x])
    
    nodes_data = []
    concept_count = 0
    for record in nodes:
        labels = record.get("labels", [])
        labels_str = ";".join(labels) if labels else ""
        is_concept = record.get("is_concept", False)
        
        created_at = record.get("created_at", "")
        updated_at = record.get("updated_at", "")
        
        # 处理时间戳
        if created_at:
            try:
                created_at = datetime.fromtimestamp(created_at / 1000).strftime("%Y-%m-%d %H:%M:%S")
            except:
                created_at = str(created_at)
        if updated_at:
            try:
                updated_at = datetime.fromtimestamp(updated_at / 1000).strftime("%Y-%m-%d %H:%M:%S")
            except:
                updated_at = str(updated_at)
        
        nodes_data.append({
            'id': record.get("id", ""),
            'name': record.get("name", ""),
            'labels': labels_str,
            'is_concept': is_concept,
            'node_role': 'concept' if is_concept else 'instance',
            'entity_type': record.get("entity_type", ""),
            'instance_of_concepts': "; ".join(instance_of_map.get(record.get("id", ""), [])),
            'concept_parent_classes': "; ".join(concept_parent_map.get(record.get("id", ""), [])),
            'concept_child_classes': "; ".join(concept_child_map.get(record.get("id", ""), [])),
            'created_at': created_at,
            'updated_at': updated_at
        })
        if is_concept:
            concept_count += 1
    
    df_nodes = pd.DataFrame(nodes_data)
    print(f"  导出了 {len(df_nodes)} 个节点 (概念: {concept_count})")
    
    # 2. 导出关系
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           s.name as source_name,
           type(r) as relation_type,
           o.id as target_id,
           o.name as target_name,
           r.is_ontology as is_ontology,
           r.evidence as evidence,
           r.source_paper_id as source_paper_id,
           r.source_paper as source_paper
    ORDER BY s.id, o.id
    """
    relationships = graph_store.query(relationships_query)
    
    rels_data = []
    ontology_rel_count = 0
    for record in relationships:
        evidence = record.get("evidence", "")
        if evidence and len(evidence) > 1000:
            evidence = evidence[:1000] + "..."
        is_ontology = record.get("is_ontology", False)
        
        rels_data.append({
            'source_id': record.get("source_id", ""),
            'source_name': record.get("source_name", ""),
            'target_id': record.get("target_id", ""),
            'target_name': record.get("target_name", ""),
            'relation_type': record.get("relation_type", ""),
            'is_ontology': is_ontology,
            'evidence': evidence,
            'source_paper_id': record.get("source_paper_id", ""),
            'source_paper': record.get("source_paper", "")
        })
        if is_ontology:
            ontology_rel_count += 1
    
    df_relationships = pd.DataFrame(rels_data)
    print(f"  导出了 {len(df_relationships)} 条关系 (本体: {ontology_rel_count})")
    
    # 3. 导出本体层次结构（新版，使用 is_ontology 标记）
    print("正在导出本体层次结构...")
    ontology_query = """
    MATCH (s)-[r]->(o)
    WHERE r.is_ontology = true 
       OR type(r) IN ['SUB_CLASS_OF', 'INSTANCE_OF', '为', '藏', '主', '生', '克', 
                      '合', '开窍于', '其华在', '走', '入', '伤', '胜', '畏', '恶']
    RETURN s.id as subject_id,
           s.name as subject_name,
           s.is_concept as subject_is_concept,
           type(r) as relation,
           o.id as object_id,
           o.name as object_name,
           o.is_concept as object_is_concept,
           r.evidence as evidence
    ORDER BY s.name, o.name
    """
    ontology_results = graph_store.query(ontology_query)
    
    ontology_data = []
    for record in ontology_results:
        evidence = record.get("evidence", "")
        if evidence and len(evidence) > 500:
            evidence = evidence[:500] + "..."
        
        ontology_data.append({
            'subject_id': record.get("subject_id", ""),
            'subject_name': record.get("subject_name", ""),
            'subject_is_concept': record.get("subject_is_concept", False),
            'relation': record.get("relation", ""),
            'object_id': record.get("object_id", ""),
            'object_name': record.get("object_name", ""),
            'object_is_concept': record.get("object_is_concept", False),
            'evidence': evidence
        })
    
    df_ontology = pd.DataFrame(ontology_data)
    print(f"  导出了 {len(df_ontology)} 条本体关系")
    
    # 4. 获取统计信息
    stats = graph_store.get_stats()
    stats_data = {
        '指标': ['节点总数', '概念节点数', '实体节点数', '关系总数', '本体关系数', '实例关系数', '导出时间'],
        '数值': [
            stats.get('nodes', 0), 
            stats.get('concept_nodes', 0),
            stats.get('entity_nodes', 0),
            stats.get('relationships', 0), 
            stats.get('ontology_relations', 0),
            stats.get('instance_relations', 0),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
    }
    df_stats = pd.DataFrame(stats_data)
    
    # 5. 写入Excel文件（多个工作表）
    print(f"正在写入Excel文件: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_nodes.to_excel(writer, sheet_name='节点', index=False)
        df_relationships.to_excel(writer, sheet_name='关系', index=False)
        df_ontology.to_excel(writer, sheet_name='本体层次', index=False)
        df_stats.to_excel(writer, sheet_name='统计信息', index=False)
    
    # 调整列宽
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    
    for sheet_name in ['节点', '关系', '本体层次']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    print(f"  工作表: 节点 ({len(df_nodes)}行, 概念:{concept_count}), 关系 ({len(df_relationships)}行, 本体:{ontology_rel_count}), 本体层次 ({len(df_ontology)}行), 统计信息")
    
    return output_file


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description="导出Neo4j图数据为多种格式（JSON、CSV、Cypher）")
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="输出文件路径或目录（默认：自动生成带时间戳的文件名）"
    )
    parser.add_argument(
        "--format",
        type=str,
        choices=["json", "simple", "cypher", "csv", "excel", "ontology", "ontology-excel"],
        default="excel",
        help="导出格式：json（完整JSON）、simple（简化JSON）、cypher（Cypher语句）、csv（CSV格式）、excel（Excel格式）、ontology（本体层次CSV）、ontology-excel（本体专用Excel）"
    )
    parser.add_argument(
        "--neo4j-uri",
        type=str,
        default=os.getenv("NEO4J_URI", "bolt://localhost:7687"),
        help="Neo4j连接URI"
    )
    parser.add_argument(
        "--neo4j-user",
        type=str,
        default=os.getenv("NEO4J_USER", "neo4j"),
        help="Neo4j用户名"
    )
    parser.add_argument(
        "--neo4j-password",
        type=str,
        default=os.getenv("NEO4J_PASSWORD", "password123"),
        help="Neo4j密码"
    )
    
    args = parser.parse_args()
    
    # 连接数据库
    try:
        print("正在连接Neo4j数据库...")
        graph_store = Neo4jGraphStore(
            uri=args.neo4j_uri,
            user=args.neo4j_user,
            password=args.neo4j_password
        )
    except Exception as e:
        print(f"连接数据库失败: {e}")
        print("\n请确保Neo4j服务正在运行:")
        print("  docker-compose up -d neo4j")
        return
    
    try:
        # 根据格式导出
        if args.format == "json":
            export_to_json(graph_store, args.output)
        elif args.format == "simple":
            export_simplified_json(graph_store, args.output)
        elif args.format == "cypher":
            export_cypher_format(graph_store, args.output)
        elif args.format == "csv":
            export_to_csv(graph_store, args.output)
        elif args.format == "excel":
            export_to_excel(graph_store, args.output)
        elif args.format == "ontology":
            export_ontology_to_csv(graph_store, args.output)
        elif args.format == "ontology-excel":
            export_ontology_excel(graph_store, args.output)
        
    except Exception as e:
        print(f"\n导出过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        graph_store.close()


if __name__ == "__main__":
    main()