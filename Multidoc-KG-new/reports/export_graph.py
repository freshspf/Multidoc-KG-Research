"""
将Neo4j图数据导出为多种格式（JSON、CSV、Excel、Cypher）
"""
import json
import csv
import os
import sys
from datetime import datetime
from typing import Dict, List, Any, Tuple

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("警告: pandas未安装，Excel导出功能不可用。请运行: pip install pandas openpyxl")

# 添加项目根目录到 Python 路径
# 这样可以从 reports/ 目录运行脚本
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from core.neo4j_store import Neo4jGraphStore


def export_to_json(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出图数据为JSON格式
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"graph_export_{timestamp}.json"
    
    print(f"\n开始导出图数据到: {output_file}")
    print("=" * 60)
    
    # 1. 导出所有节点
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    RETURN n.id as id, 
           n.name as name, 
           labels(n) as labels,
           properties(n) as properties
    """
    nodes = graph_store.query(nodes_query)
    nodes_list = []
    for record in nodes:
        node_data = {
            "id": record.get("id"),
            "name": record.get("name"),
            "labels": record.get("labels", []),
            "properties": {k: v for k, v in (record.get("properties") or {}).items() 
                          if k not in ["id", "name"]}  # 避免重复
        }
        nodes_list.append(node_data)
    
    print(f"  导出了 {len(nodes_list)} 个节点")
    
    # 2. 导出所有关系
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           s.name as source_name,
           type(r) as relation_type,
           o.id as target_id,
           o.name as target_name,
           properties(r) as properties
    """
    relationships = graph_store.query(relationships_query)
    relationships_list = []
    for record in relationships:
        rel_data = {
            "source": {
                "id": record.get("source_id"),
                "name": record.get("source_name")
            },
            "target": {
                "id": record.get("target_id"),
                "name": record.get("target_name")
            },
            "relation": record.get("relation_type"),
            "properties": {k: v for k, v in (record.get("properties") or {}).items()}
        }
        relationships_list.append(rel_data)
    
    print(f"  导出了 {len(relationships_list)} 条关系")
    
    # 3. 获取统计信息
    print("正在获取统计信息...")
    stats = graph_store.get_stats()
    
    # 4. 组装导出数据
    export_data = {
        "metadata": {
            "export_time": datetime.now().isoformat(),
            "nodes_count": len(nodes_list),
            "relationships_count": len(relationships_list),
            "database_stats": stats
        },
        "nodes": nodes_list,
        "relationships": relationships_list
    }
    
    # 5. 写入JSON文件
    print(f"正在写入文件: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    print(f"  节点数: {len(nodes_list)}")
    print(f"  关系数: {len(relationships_list)}")
    
    return output_file


def export_simplified_json(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出简化版JSON（仅包含基本信息，文件更小）
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"graph_export_simple_{timestamp}.json"
    
    print(f"\n开始导出简化版图数据到: {output_file}")
    print("=" * 60)
    
    # 导出节点（仅ID和名称）
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    RETURN DISTINCT n.id as id, n.name as name
    """
    nodes = graph_store.query(nodes_query)
    nodes_dict = {record.get("id"): record.get("name") 
                  for record in nodes if record.get("id")}
    
    print(f"  导出了 {len(nodes_dict)} 个节点")
    
    # 导出关系（简化版）
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           type(r) as relation_type,
           o.id as target_id,
           r.evidence as evidence,
           r.source_paper as source_paper
    """
    relationships = graph_store.query(relationships_query)
    relationships_list = []
    for record in relationships:
        rel_data = {
            "source_id": record.get("source_id"),
            "target_id": record.get("target_id"),
            "relation": record.get("relation_type"),
            "evidence": record.get("evidence", "")[:200] if record.get("evidence") else "",  # 限制长度
            "source_paper": record.get("source_paper", "")
        }
        relationships_list.append(rel_data)
    
    print(f"  导出了 {len(relationships_list)} 条关系")
    
    # 组装数据
    export_data = {
        "metadata": {
            "export_time": datetime.now().isoformat(),
            "nodes_count": len(nodes_dict),
            "relationships_count": len(relationships_list)
        },
        "nodes": nodes_dict,
        "relationships": relationships_list
    }
    
    # 写入文件
    print(f"正在写入文件: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    
    return output_file


def export_cypher_format(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出为Cypher格式（可以用于重新导入）
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"graph_export_{timestamp}.cypher"
    
    print(f"\n开始导出Cypher格式到: {output_file}")
    print("=" * 60)
    
    cypher_statements = []
    
    # 导出节点
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    RETURN n.id as id, n.name as name, labels(n) as labels
    """
    nodes = graph_store.query(nodes_query)
    node_count = 0
    
    for record in nodes:
        node_id = record.get("id", "").replace("'", "\\'")
        node_name = record.get("name", "").replace("'", "\\'")
        labels = record.get("labels", ["Entity"])
        label_str = ":".join(labels)
        
        cypher = f"CREATE (n:{label_str} {{id: '{node_id}', name: '{node_name}'}});"
        cypher_statements.append(cypher)
        node_count += 1
    
    print(f"  导出了 {node_count} 个节点")
    
    # 导出关系
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           type(r) as relation_type,
           o.id as target_id,
           r.evidence as evidence,
           r.source_paper as source_paper
    """
    relationships = graph_store.query(relationships_query)
    rel_count = 0
    
    for record in relationships:
        source_id = record.get("source_id", "").replace("'", "\\'")
        target_id = record.get("target_id", "").replace("'", "\\'")
        rel_type = record.get("relation_type", "RELATES_TO")
        evidence = (record.get("evidence", "") or "").replace("'", "\\'")[:100]
        source_paper = (record.get("source_paper", "") or "").replace("'", "\\'")
        
        cypher = f"""MATCH (s {{id: '{source_id}'}}), (o {{id: '{target_id}'}})
CREATE (s)-[r:{rel_type} {{evidence: '{evidence}', source_paper: '{source_paper}'}}]->(o);"""
        cypher_statements.append(cypher)
        rel_count += 1
    
    print(f"  导出了 {rel_count} 条关系")
    
    # 写入文件
    print(f"正在写入文件: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("// Neo4j Graph Export\n")
        f.write(f"// Export time: {datetime.now().isoformat()}\n")
        f.write(f"// Nodes: {node_count}, Relationships: {rel_count}\n\n")
        f.write("\n".join(cypher_statements))
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    
    return output_file


def export_to_csv(graph_store: Neo4jGraphStore, output_dir: str = None):
    """
    导出图数据为CSV格式（节点和关系分别导出为两个CSV文件）
    
    Args:
        graph_store: Neo4j图存储实例
        output_dir: 输出目录（可选，默认为当前目录）
    """
    if output_dir is None:
        output_dir = "."
    else:
        os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nodes_file = os.path.join(output_dir, f"nodes_{timestamp}.csv")
    relationships_file = os.path.join(output_dir, f"relationships_{timestamp}.csv")
    
    print(f"\n开始导出CSV格式图数据")
    print("=" * 60)
    
    # 1. 导出节点
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    WHERE n:Entity OR n:Ontology
    RETURN n.id as id, 
           n.name as name, 
           labels(n) as labels,
           n.entity_type as entity_type,
           n.created_at as created_at,
           n.updated_at as updated_at
    ORDER BY n.id
    """
    nodes = graph_store.query(nodes_query)
    
    node_count = 0
    with open(nodes_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        # 写入表头
        writer.writerow([
            'id', 'name', 'labels', 'entity_type', 
            'created_at', 'updated_at'
        ])
        
        for record in nodes:
            node_id = record.get("id", "")
            node_name = record.get("name", "")
            labels = record.get("labels", [])
            labels_str = ";".join(labels) if labels else ""
            entity_type = record.get("entity_type", "")
            created_at = record.get("created_at", "")
            updated_at = record.get("updated_at", "")
            
            # 处理时间戳
            if created_at:
                try:
                    created_at = datetime.fromtimestamp(created_at / 1000).isoformat()
                except:
                    created_at = str(created_at)
            if updated_at:
                try:
                    updated_at = datetime.fromtimestamp(updated_at / 1000).isoformat()
                except:
                    updated_at = str(updated_at)
            
            writer.writerow([
                node_id, node_name, labels_str, entity_type,
                created_at, updated_at
            ])
            node_count += 1
    
    print(f"  导出了 {node_count} 个节点到: {nodes_file}")
    
    # 2. 导出关系
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           s.name as source_name,
           type(r) as relation_type,
           o.id as target_id,
           o.name as target_name,
           r.evidence as evidence,
           r.source_paper_id as source_paper_id,
           r.source_paper as source_paper
    ORDER BY s.id, o.id
    """
    relationships = graph_store.query(relationships_query)
    
    rel_count = 0
    with open(relationships_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        # 写入表头
        writer.writerow([
            'source_id', 'source_name', 'target_id', 'target_name',
            'relation_type', 'evidence', 'source_paper_id', 'source_paper'
        ])
        
        for record in relationships:
            source_id = record.get("source_id", "")
            source_name = record.get("source_name", "")
            target_id = record.get("target_id", "")
            target_name = record.get("target_name", "")
            relation_type = record.get("relation_type", "")
            evidence = record.get("evidence", "")
            source_paper_id = record.get("source_paper_id", "")
            source_paper = record.get("source_paper", "")
            
            # 限制 evidence 长度（CSV 中避免过长）
            if evidence and len(evidence) > 500:
                evidence = evidence[:500] + "..."
            
            writer.writerow([
                source_id, source_name, target_id, target_name,
                relation_type, evidence, source_paper_id, source_paper
            ])
            rel_count += 1
    
    print(f"  导出了 {rel_count} 条关系到: {relationships_file}")
    
    # 3. 计算文件大小
    nodes_size = os.path.getsize(nodes_file) / (1024 * 1024)  # MB
    rels_size = os.path.getsize(relationships_file) / (1024 * 1024)  # MB
    
    print(f"\n导出完成！")
    print(f"  节点文件: {nodes_file} ({nodes_size:.2f} MB)")
    print(f"  关系文件: {relationships_file} ({rels_size:.2f} MB)")
    print(f"  节点数: {node_count}")
    print(f"  关系数: {rel_count}")
    
    return nodes_file, relationships_file


def export_ontology_to_csv(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出本体层次结构为CSV（实体-本体类别的IS_A关系）
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"ontology_hierarchy_{timestamp}.csv"
    
    print(f"\n开始导出本体层次结构到: {output_file}")
    print("=" * 60)
    
    # 查询实体-本体关系
    # 兼容两种图谱写法：
    # 1) 旧版：(:Entity)-[:IS_A]->(:Ontology)
    # 2) 新版：类型关系以普通边写入（如 rdf:type 被清洗为 RDFTYPE），且“本体类”也是 :Entity 节点
    query_primary = """
    MATCH (e:Entity)-[:IS_A]->(o:Ontology)
    RETURN e.id as entity_id,
           e.name as entity_name,
           e.entity_type as entity_type,
           o.name as ontology_class
    ORDER BY o.name, e.name
    """
    results = graph_store.query(query_primary)

    if not results:
        print("  未发现 (:Entity)-[:IS_A]->(:Ontology) 结构，尝试从类型关系推断本体层次（IS_A / RDFTYPE / RDF_TYPE）...")
        query_fallback = """
        MATCH (e:Entity)-[r]->(o:Entity)
        WHERE type(r) IN ['IS_A', 'RDFTYPE', 'RDF_TYPE']
        RETURN e.id as entity_id,
               e.name as entity_name,
               e.entity_type as entity_type,
               coalesce(o.name, o.id) as ontology_class
        ORDER BY ontology_class, entity_name
        """
        results = graph_store.query(query_fallback)
    
    count = 0
    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        # 写入表头
        writer.writerow(['entity_id', 'entity_name', 'entity_type', 'ontology_class'])
        
        for record in results:
            writer.writerow([
                record.get("entity_id", ""),
                record.get("entity_name", ""),
                record.get("entity_type", ""),
                record.get("ontology_class", "")
            ])
            count += 1
    
    file_size = os.path.getsize(output_file) / 1024  # KB
    print(f"  导出了 {count} 条实体-本体关系")
    print(f"  文件: {output_file} ({file_size:.2f} KB)")
    
    return output_file


def export_to_excel(graph_store: Neo4jGraphStore, output_file: str = None):
    """
    导出图数据为Excel格式（包含多个工作表：节点、关系、本体层次）
    
    Args:
        graph_store: Neo4j图存储实例
        output_file: 输出文件路径（可选）
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("pandas和openpyxl未安装，无法导出Excel。请运行: pip install pandas openpyxl")
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"graph_export_{timestamp}.xlsx"
    
    print(f"\n开始导出Excel格式图数据到: {output_file}")
    print("=" * 60)
    
    # 1. 导出节点
    print("正在导出节点...")
    nodes_query = """
    MATCH (n)
    WHERE n:Entity OR n:Ontology
    RETURN n.id as id, 
           n.name as name, 
           labels(n) as labels,
           n.entity_type as entity_type,
           n.created_at as created_at,
           n.updated_at as updated_at
    ORDER BY n.id
    """
    nodes = graph_store.query(nodes_query)
    
    nodes_data = []
    for record in nodes:
        labels = record.get("labels", [])
        labels_str = ";".join(labels) if labels else ""
        
        created_at = record.get("created_at", "")
        updated_at = record.get("updated_at", "")
        
        # 处理时间戳
        if created_at:
            try:
                created_at = datetime.fromtimestamp(created_at / 1000).strftime("%Y-%m-%d %H:%M:%S")
            except:
                created_at = str(created_at)
        if updated_at:
            try:
                updated_at = datetime.fromtimestamp(updated_at / 1000).strftime("%Y-%m-%d %H:%M:%S")
            except:
                updated_at = str(updated_at)
        
        nodes_data.append({
            'id': record.get("id", ""),
            'name': record.get("name", ""),
            'labels': labels_str,
            'entity_type': record.get("entity_type", ""),
            'created_at': created_at,
            'updated_at': updated_at
        })
    
    df_nodes = pd.DataFrame(nodes_data)
    print(f"  导出了 {len(df_nodes)} 个节点")
    
    # 2. 导出关系
    print("正在导出关系...")
    relationships_query = """
    MATCH (s)-[r]->(o)
    RETURN s.id as source_id,
           s.name as source_name,
           type(r) as relation_type,
           o.id as target_id,
           o.name as target_name,
           r.evidence as evidence,
           r.source_paper_id as source_paper_id,
           r.source_paper as source_paper
    ORDER BY s.id, o.id
    """
    relationships = graph_store.query(relationships_query)
    
    rels_data = []
    for record in relationships:
        evidence = record.get("evidence", "")
        if evidence and len(evidence) > 1000:
            evidence = evidence[:1000] + "..."
        
        rels_data.append({
            'source_id': record.get("source_id", ""),
            'source_name': record.get("source_name", ""),
            'target_id': record.get("target_id", ""),
            'target_name': record.get("target_name", ""),
            'relation_type': record.get("relation_type", ""),
            'evidence': evidence,
            'source_paper_id': record.get("source_paper_id", ""),
            'source_paper': record.get("source_paper", "")
        })
    
    df_relationships = pd.DataFrame(rels_data)
    print(f"  导出了 {len(df_relationships)} 条关系")
    
    # 3. 导出本体层次结构
    print("正在导出本体层次结构...")
    # 兼容两种图谱写法（同 export_ontology_to_csv）：
    ontology_query_primary = """
    MATCH (e:Entity)-[:IS_A]->(o:Ontology)
    RETURN e.id as entity_id,
           e.name as entity_name,
           e.entity_type as entity_type,
           o.name as ontology_class
    ORDER BY o.name, e.name
    """
    ontology_results = graph_store.query(ontology_query_primary)

    if not ontology_results:
        print("  未发现 (:Entity)-[:IS_A]->(:Ontology) 结构，尝试从类型关系推断本体层次（IS_A / RDFTYPE / RDF_TYPE）...")
        ontology_query_fallback = """
        MATCH (e:Entity)-[r]->(o:Entity)
        WHERE type(r) IN ['IS_A', 'RDFTYPE', 'RDF_TYPE']
        RETURN e.id as entity_id,
               e.name as entity_name,
               e.entity_type as entity_type,
               coalesce(o.name, o.id) as ontology_class
        ORDER BY ontology_class, entity_name
        """
        ontology_results = graph_store.query(ontology_query_fallback)
    
    ontology_data = []
    for record in ontology_results:
        ontology_data.append({
            'entity_id': record.get("entity_id", ""),
            'entity_name': record.get("entity_name", ""),
            'entity_type': record.get("entity_type", ""),
            'ontology_class': record.get("ontology_class", "")
        })
    
    df_ontology = pd.DataFrame(ontology_data)
    print(f"  导出了 {len(df_ontology)} 条本体关系")
    
    # 4. 获取统计信息
    stats = graph_store.get_stats()
    stats_data = {
        '指标': ['节点总数', '关系总数', '导出时间'],
        '数值': [stats.get('nodes', 0), stats.get('relationships', 0), datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    }
    df_stats = pd.DataFrame(stats_data)
    
    # 5. 写入Excel文件（多个工作表）
    print(f"正在写入Excel文件: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_nodes.to_excel(writer, sheet_name='节点', index=False)
        df_relationships.to_excel(writer, sheet_name='关系', index=False)
        df_ontology.to_excel(writer, sheet_name='本体层次', index=False)
        df_stats.to_excel(writer, sheet_name='统计信息', index=False)
    
    # 调整列宽（仅对前三个工作表）
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    
    for sheet_name in ['节点', '关系', '本体层次']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    
    file_size = os.path.getsize(output_file) / (1024 * 1024)  # MB
    print(f"\n导出完成！")
    print(f"  文件: {output_file}")
    print(f"  文件大小: {file_size:.2f} MB")
    print(f"  工作表: 节点 ({len(df_nodes)}行), 关系 ({len(df_relationships)}行), 本体层次 ({len(df_ontology)}行), 统计信息")
    
    return output_file


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description="导出Neo4j图数据为多种格式（JSON、CSV、Cypher）")
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="输出文件路径或目录（默认：自动生成带时间戳的文件名）"
    )
    parser.add_argument(
        "--format",
        type=str,
        choices=["json", "simple", "cypher", "csv", "excel", "ontology"],
        default="json",
        help="导出格式：json（完整JSON）、simple（简化JSON）、cypher（Cypher语句）、csv（CSV格式）、excel（Excel格式）、ontology（本体层次CSV）"
    )
    parser.add_argument(
        "--neo4j-uri",
        type=str,
        default=os.getenv("NEO4J_URI", "bolt://localhost:7687"),
        help="Neo4j连接URI"
    )
    parser.add_argument(
        "--neo4j-user",
        type=str,
        default=os.getenv("NEO4J_USER", "neo4j"),
        help="Neo4j用户名"
    )
    parser.add_argument(
        "--neo4j-password",
        type=str,
        default=os.getenv("NEO4J_PASSWORD", "password123"),
        help="Neo4j密码"
    )
    
    args = parser.parse_args()
    
    # 连接数据库
    try:
        print("正在连接Neo4j数据库...")
        graph_store = Neo4jGraphStore(
            uri=args.neo4j_uri,
            user=args.neo4j_user,
            password=args.neo4j_password
        )
    except Exception as e:
        print(f"连接数据库失败: {e}")
        print("\n请确保Neo4j服务正在运行:")
        print("  docker-compose up -d neo4j")
        return
    
    try:
        # 根据格式导出
        if args.format == "json":
            export_to_json(graph_store, args.output)
        elif args.format == "simple":
            export_simplified_json(graph_store, args.output)
        elif args.format == "cypher":
            export_cypher_format(graph_store, args.output)
        elif args.format == "csv":
            export_to_csv(graph_store, args.output)
        elif args.format == "excel":
            export_to_excel(graph_store, args.output)
        elif args.format == "ontology":
            export_ontology_to_csv(graph_store, args.output)
        
    except Exception as e:
        print(f"\n导出过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        graph_store.close()


if __name__ == "__main__":
    main()
